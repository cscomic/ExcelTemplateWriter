#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Author   : cscomic
# @Time     : 2022/5/2 16:17
# @File     : excel_template_writer.py
# @Project  : ExcelTemplateWriter
# @Version  : 0.0.1
# @License  : 
# @Desc     :
from typing import NoReturn

from pandas import DataFrame, isnull
from openpyxl.utils.dataframe import dataframe_to_rows
from src.exceltemplatewriter.extended_openpyxl_writer import load_workbook, save_workbook


def write_dataframe_to_excel_template(data: DataFrame,
                                      excel_template: str,
                                      sheet_name: str = None,
                                      output_file: str = None,
                                      ignore_empty_cell: bool = True) -> NoReturn:
    """
    Write data in Dataframe type into excel file with keeping format and formula.
    Args:
        data (pandas.Dataframe): data we want to write into file
        excel_template (str): name and path of your excel template file
        sheet_name (str): which sheet we want to output our data, by default is 'Sheet1'
        output_file (str): name and path of output data file, if this value is None, the data will write to template file directly.
                            strongly recommend not to do so, or your data in template may lost
        ignore_empty_cell (bool): if the data is empty, we will skip to write the cell, we can use this make some formula in template
    """
    workbook = load_workbook(excel_template, keep_links=False)

    if sheet_name is None:
        sheet_name = "Sheet1"
    sheet = workbook[sheet_name]
    rows = dataframe_to_rows(data, index=False, header=True)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            if not ignore_empty_cell \
                    or (value is not None
                        and not isnull(value)
                        and value != ""):
                sheet.cell(row=r_idx, column=c_idx, value=value)

    if output_file is None:
        output_file = excel_template

    save_workbook(workbook, output_file)
